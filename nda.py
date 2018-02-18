#!/usr/bin/env python

'''
---AUTHOR---
Name: Matt Cross
Email: routeallthings@gmail.com

---PREREQ---
INSTALL netmiko (pip install netmiko)
INSTALL textfsm (pip install textfsm)
INSTALL openpyxl (pip install openpyxl)
INSTALL fileinput (pip install fileinput)
INSTALL xlhelper (python -m pip install git+git://github.com/routeallthings/xlhelper.git)
'''

#Module Imports (Native)
import re
import getpass
import os
import unicodedata
import csv
import threading
import time
import sys
import json
import logging
import datetime
from decimal import *

# FIXES

paramiko_logger = logging.getLogger('paramiko.transport')
if not paramiko_logger.handlers:
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(
        logging.Formatter('%(asctime)s | %(levelname)-8s| PARAMIKO: '
                      '%(lineno)03d@%(module)-10s| %(message)s')
    )
paramiko_logger.addHandler(console_handler)

#Module Imports (Non-Native)

# Used for pulling information for MAC table
try:
	import requests
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('request module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in requestsinstallstatus.lower():
		os.system('python -m pip install requests')
		import requests
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of requests. Please install manually and retry"
		sys.exit()

# Used because its the best way to connect into the switches to pull out information (besides SNMP)
try:
	import netmiko
	from netmiko import ConnectHandler
except ImportError:
	netmikoinstallstatus = fullpath = raw_input ('Netmiko module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in netmikoinstallstatus.lower():
		os.system('python -m pip install netmiko')
		import netmiko
		from netmiko import ConnectHandler
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of netmiko. Please install manually and retry"
		sys.exit()

# Used to parse through SSH output to get relevant data into tables for comparison and output
try:
	import textfsm
except ImportError:
	textfsminstallstatus = fullpath = raw_input ('textfsm module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in textfsminstallstatus.lower():
		os.system('python -m pip install textfsm')
		import textfsm
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of textfsm. Please install manually and retry"
		sys.exit()

# XLSX import
try:
	from openpyxl import load_workbook
	from openpyxl import workbook
	from openpyxl import Workbook
except ImportError:
	openpyxlinstallstatus = fullpath = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in openpyxlinstallstatus.lower():
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
		from openpyxl import workbook
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of Pandas. Please install manually and retry'
		sys.exit()

# XLSX import
try:
	import fileinput
except ImportError:
	fileinputnstallstatus = fullpath = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in fileinputnstallstatus.lower():
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		sys.exit()

# TextFSM download function
try:
	import urllib
except ImportError:
	urllibinstallstatus = fullpath = raw_input ('urllib module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in urllibinstallstatus.lower():
		os.system('python -m pip install urllib')
		import urllib
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of urllib. Please install manually and retry"
		sys.exit()

# Darth-Veitcher Module https://github.com/darth-veitcher/xlhelper
# XLSX Import
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
try:
	import xlhelper
except ImportError:
	xlhelperinstallstatus = fullpath = raw_input ('xlhelper module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in xlhelperinstallstatus.lower():
		os.system('python -m pip install git+git://github.com/routeallthings/xlhelper.git')
		import xlhelper
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of xlhelper. Please install manually and retry'
		sys.exit()
		

# MNETSUITE (https://github.com/MJL85/mnet/tree/master/mnetsuite)---- THIS HAS BEEN MODIFIED IN ROUTEALLTHINGS FORK
# CDP/LLDP Discovery and Mapping method

try:
	import mnetsuite
except ImportError:
	mnetinstallstatus = fullpath = raw_input ('mnetsuite module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in mnetinstallstatus.lower():
		os.system('python -m pip install git+git://github.com/routeallthings/mnet.git')
		import mnetsuite
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of mnetsuite. Please install manually and retry'
		sys.exit()
		
# Graphviz

try:
	from graphviz import Digraph
except ImportError:
	graphvizinstallstatus = fullpath = raw_input ('graphciz module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in graphvizinstallstatus.lower():
		os.system('python -m pip install graphviz')
		from graphviz import Digraph
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of graphviz. Please install manually and retry'
		sys.exit()

'''Global Variables'''
ipv4_address = re.compile('((25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?).){3}(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)')

#Creation of SSH commands
showrun = "show running-config"
showstart = "show startup-config"
showver = "show version"
showinv = "show inventory"
showcdp = "show cdp neighbors detail"
showlldp = "show lldp neighbors detail"
showpowerinline = "show power inline"
showstackpower = "show stack-power detail"
showswitch = "show switch detail"
showdhcpsnooping = "show ip dhcp snooping"
showvlan = "show vlan"
showtrunk = "show interface trunk"
showspanning = "show spanning-tree"
showspanningblock = "show spanning-tree blockedports"
showinterfacestat = "show interface status"
showipintbr = "show ip interface brief"
showeigrpnei = "show ip eigrp neighbors"
showeigrptop = "show ip eigrp topology all"
showospfnei = "show ip ospf neighbors"
showospfdata = "show ip ospf database"
showbgpnei = "show ip bgp neighbors"
showbgptable = "show ip bgp"
showiproute = "show ip route"
showvrf = "show vrf"
showtemp = "show env temperature status"
showippimnei = "show ip pim neighbor"
showmroute = "show ip mroute"
showigmpsnoop = "show ip igmp snooping"
showipigmpmember = "show ip igmp membership"
showinterface = "show interface"
showinterfacet = "show interface transceiver"
showiparp = "show ip arp"
showmacaddress = "show mac address-table"
showmacaddress_older = "show mac-address-table"
showlocation = "show run | i ^snmp-server location"

# Device Match Lists
ciscoxelist = '3650 3850 9300 9400 9500 4500 4431 4451 4321 4331 4351 asr'
ciscoioslist = '3750 2960 3560 6500 2801 2811 2821 2851 2911 2921 2951 2901 3825 3845'
cisconxoslist = '7700 7000 Nexus'

# Device Match Lists Convert
ciscoxelist = ciscoxelist.split(' ')
ciscoioslist = ciscoioslist.split(' ')
cisconxoslist = cisconxoslist.split(' ')

# Device Type Lists
switchlist = 'Catalyst 9300 9400 9500 2950 2960 3550 3560 3750 3650 3850 4500 6500 WS-C Nexus NXOS'
routerlist = '2801 2811 2821 2851 2901 2911 2921 2951 3825 3845 4300 4400 4321 4331 4351 4431 4451 asr csv'
fwlist = 'ASA'

# Device Type Convert
switchlist = switchlist.split(' ')
routerlist = routerlist.split(' ')
fwlist = fwlist.split(' ')

# Cisco Version Number Regex
ciscoverreg = '1[256]\.[1-9]\(.*\).*'

# Create empty lists for script use
healthchecklist = []
tempfilelist = []
cdpdevicecomplete = []
cdpdevicediscovery = []
sshdevices = []
usernamelist = []

# Create empty lists for export use
fullinventorylist = []
ipmactablelist = []
mactablelist = []
iparptablelist = []
l2interfacelist = []
l3interfacelist = []
vlanlist = []
poeinterfacelist = []

# URLs to use
maclookupurl = 'http://macvendors.co/api/%s'
maclookupdburl = "https://linuxnet.ca/ieee/oui.txt"

'''Global Variable Questions'''
print ''
print 'Network Documention Automation'
print '##########################################################'
print 'The purpose of this tool is to pull information from the'
print 'network via CDP/LLDP discovery or manual entry in XLSX.'
print 'Please fill in the configuration tab on the templated'
print 'XLSX sheet, along with all the data that you want to test.'
print '##########################################################'
print ''
print '----Questions that need answering----'
excelfilelocation = raw_input('File to load the excel data from (e.g. C:/Python27/nda-config.xlsx):')
if excelfilelocation == '':
	excelfilelocation = 'C:/Python27/nda-config.xlsx'
excelfilelocation = excelfilelocation.replace('"', '')
# Load Configuration Variables
configdict = {}
for configvariables in xlhelper.sheet_to_dict(excelfilelocation,'Config'):
	try:
		configvar = configvariables.get('Variable').encode('utf-8')
		configval = configvariables.get('Value').encode('utf-8')
	except:
		configvar = configvariables.get('Variable')
		configval = configvariables.get('Value')
	configdict[configvar] = configval
# Username Variables/Questions
for usernames in xlhelper.sheet_to_dict(excelfilelocation,'Username'):
	usernamedict = {}
	try:
		usernamev = usernames.get('Username').encode('utf-8')
		passwordv = usernames.get('Password').encode('utf-8')
		enablev = usernames.get('Enable Password').encode('utf-8')
	except:
		usernamev = usernames.get('Username')
		passwordv = usernames.get('Password')
		enablev = usernames.get('Enable Password')
	try:
		usernamedict['sshusername'] = usernamev
		usernamedict['sshpassword'] = passwordv
		usernamedict['enablesecret'] = enablev
	except:
		pass
	usernamelist.append(usernamedict)
if usernamelist == []:
	sshusername = raw_input('What is the username you will use to login to the equipment?:')
	sshpassword = getpass.getpass('What is the password you will use to login to the equipment?:')
	enablesecret = getpass.getpass('What is the enable password you will use to access the device?:')
# Rest of the Config Variables
exportlocation = configdict.get('ExportLocation')
if exportlocation == '':
	exportlocation = r'C:\OutputFolder'
#### DEVICE DISCOVERY #####
devicediscoveryv = configdict.get('DeviceDiscovery')
if devicediscoveryv == None:
	devicediscoveryv = 0
devicediscoveryseedv = configdict.get('DeviceDiscoverySeed')
devicediscoverysshv = configdict.get('DeviceDiscoverySSH')
if devicediscoverysshv == None:
	devicediscoverysshv = 0
devicediscoveryseedv = configdict.get('DeviceDiscoverySeed')
if devicediscoveryseedv == None:
	print 'No CDP seed router, not doing CDP/LLDP discovery'
	devicediscoveryseedv = 'NA'
devicediscoverydepthv = configdict.get('DeviceDiscoveryDepth')
if devicediscoverydepthv == None:
	devicediscoverydepthv = 0
devicediscoverysshtypev = configdict.get('DeviceDiscoverySSHType')
devicediscoverysshtypev = devicediscoverysshtypev.lower()
if devicediscoverysshtypev == None:
	devicediscoverysshtypev = 'cisco_ios'
if 'nxos' in devicediscoverysshtypev.lower() or 'ios' in devicediscoverysshtypev.lower() or 'xe' in devicediscoverysshtypev.lower():
	devicediscoverysshtypev = devicediscoverysshtypev.encode('utf-8')
	devicediscoverysshtypev = 'cisco_' + devicediscoverysshtypev
#### MNET CDP DISCOVERY ####
devicediscoverymapv = configdict.get('DeviceDiscoveryMap')
if devicediscoverymapv == None:
	devicediscoverymapv = 0
devicediscoverymaptitlev = configdict.get('DeviceDiscoveryMapTitle')
if devicediscoverymaptitlev == None:
	devicediscoverymaptitlev = 'Network Topology'
if devicediscoverymapv == 1:
	import pydot
	pydottest = pydot.find_graphviz()
	if not 'neato' in pydottest:
		print 'Could not find graphviz. Please make sure the PATH variable is set in windows to the correct location, and that the product is installed'
mnetvar = {}
mnetsnmp = []
mnetdomains = []
mnetexclude = []
mnetsubnets = []
mnetgraph = {}
# MNET SNMP
for snmpvariables in xlhelper.sheet_to_dict(excelfilelocation,'SNMP'):
	snmpdict = {}
	try:
		snmpcommunityv = snmpvariables.get('SNMP Community').encode('utf-8')
		snmpversionv = snmpvariables.get('Version').encode('utf-8')
	except:
		snmpcommunityv = snmpvariables.get('SNMP Community')
		snmpversionv = snmpvariables.get('Version')
	try:
		snmpdict['community'] = snmpcommunityv
		snmpdict['ver'] = snmpversionv
		mnetsnmp.append(snmpdict)
	except:
		devicediscoverysnmp = 'NA'
# MNET DOMAINS
devicediscoverydomains = configdict.get('DeviceDiscoveryDomains')
if devicediscoverydomains == None:
	devicediscoverydomains = 'NA'
if ',' in devicediscoverydomains:
	devicediscoverydomains = devicediscoverydomains.split(',')
	for device in devicediscoverydomains:
		mnetdomains.append(device)
else:
	mnetdomains.append(devicediscoverydomains)
# MNET Exclude
devicediscoveryexcludedsubnets = configdict.get('DeviceDiscoveryExcludedSubnets')
if devicediscoveryexcludedsubnets == None:
	devicediscoveryexcludedsubnets = '255.255.255.255/32'
if ',' in devicediscoveryexcludedsubnets:
	devicediscoveryexcludedsubnets = devicediscoveryexcludedsubnets.split(',')
	for device in devicediscoveryexcludedsubnets:
		mnetexclude.append(device)
else:
	mnetexclude.append(devicediscoveryexcludedsubnets)
# MNET Subnets
devicediscoveryincludedsubnets = configdict.get('DeviceDiscoveryIncludedSubnets')
if devicediscoveryincludedsubnets == None:
	devicediscoveryincludedsubnets = '10.0.0.0/8,192.168.0.0/16,172.16.0.0/12'
if ',' in devicediscoveryincludedsubnets:
	devicediscoveryincludedsubnets = devicediscoveryincludedsubnets.split(',')
	for device in devicediscoveryincludedsubnets:
		mnetsubnets.append(device)
else:
	mnetsubnets.append(devicediscoveryincludedsubnets)
# MNET Graph
mnetgraph['node_text_size'] = 10
mnetgraph['link_text_size'] = 9
mnetgraph['title_text_size'] = 15
mnetgraph['include_svi'] = 1
mnetgraph['include_lo'] = 1
mnetgraph['include_serials'] = 1
mnetgraph['get_stack_members'] = 1
mnetgraph['get_vss_members'] = 1
mnetgraph['expand_stackwise'] = 0
mnetgraph['expand_vss'] = 0
mnetgraph['expand_lag'] = 0
# MNET Full
mnetvar['snmp'] = mnetsnmp
mnetvar['domains'] = mnetdomains
mnetvar['exclude'] = mnetexclude
mnetvar['subnets'] = mnetsubnets
mnetvar['graph'] = mnetgraph
mnetvar['exclude_hosts'] = []
mnetfile = 'nda-mnetvar.conf'
mnetcat = 'nda-mnetcat.csv'
with open(mnetfile,'w') as jsonfile:
	json.dump(mnetvar, jsonfile)
tempfilelist.append(mnetfile)
tempfilelist.append(mnetcat)
##### MNET END #####
# Normal Config Vars	
configurationv = configdict.get('Configuration')
if configurationv == None:
	configurationv = 'True'
healthcheckv = configdict.get('HealthCheck')
if healthcheckv == None:
	healthcheckv = 'True'
powerbudget = configdict.get('PowerBudget')
if powerbudget == None:
	powerbudget = 'Y'
dhcpsnooping = configdict.get('DHCPSnooping')
if dhcpsnooping == None:
	dhcpsnooping = 'Y'
temperature = configdict.get('Temperature')
if temperature == None:
	temperature = 'Y'
# End of Config Variables
# Start of Functions

def DEF_REMOVEPREFIX(text, prefix):
    if text.startswith(prefix):
        return text[len(prefix):]
    return text

def DEF_STARTALLTESTS(sshdevice):
	if configurationv == 1:
		DEF_GATHERDATA(sshdevice)
	if healthcheckv == 1:
		DEF_HEALTHCHECK(sshdevice)

def DEF_WRITEOUTPUT(sshcommand,sshresult,sshdevicehostname,outputfolder):
	sshcommandfile = sshcommand.replace(' ','')
	sshcommandfile = sshcommandfile.replace('-','')
	outputfile = outputfolder + '\\' + sshdevicehostname + '_' + sshcommandfile + '.txt'
	f = open(outputfile,'w')
	f.write(sshresult)
	f.close()

##### FUNCTIONS #####
def DEF_GATHERDATA(sshdevice):
	sshdeviceip = sshdevice.get('Device IPs').encode('utf-8')
	sshdevicevendor = sshdevice.get('Vendor').encode('utf-8')
	sshdevicetype = sshdevice.get('Type').encode('utf-8')
	sshdevicetype = sshdevicevendor.lower() + "_" + sshdevicetype.lower()
	# Device Type Assignment
	deviceswitch = 0
	devicerouter = 0
	deviceasa = 0
	#Start Connection
	try:
		for username in usernamelist:
			try:
				sshusername = username.get('sshusername').encode('utf-8')
				sshpassword = username.get('sshpassword').encode('utf-8')
				enablesecret = username.get('enablesecret').encode('utf-8')
			except:
				sshusername = username.get('sshusername')
				sshpassword = username.get('sshpassword')
				enablesecret = username.get('enablesecret')
			try:
				sshnet_connect = ConnectHandler(device_type=sshdevicetype, ip=sshdeviceip, username=sshusername, password=sshpassword, secret=enablesecret)
				break
			except:
				pass
		if not sshnet_connect:
			thread.exit()
		sshdevicehostname = sshnet_connect.find_prompt()
		sshdevicehostname = sshdevicehostname.strip('#')
		if '>' in sshdevicehostname:
			sshnet_connect.enable()
			sshdevicehostname = sshdevicehostname.strip('>')
			sshdevicehostname = sshnet_connect.find_prompt()
			sshdevicehostname = sshdevicehostname.strip('#')
		print 'Successfully connected to ' + sshdevicehostname
		print 'Gathering data from ' + sshdevicehostname
		#Create output folder if none exists
		outputfolder = exportlocation + '\\' + sshdevicehostname
		if not os.path.exists(outputfolder):
			os.makedirs(outputfolder)
		#################################################################
		################### DOWNLOAD TEMPLATES START ####################
		#################################################################
		# Show Inventory
		if "cisco_ios" in sshdevicetype.lower():
			fsmshowinvurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_inventory.template"
		if "cisco_xe" in sshdevicetype.lower():
			fsmshowinvurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_inventory.template"
		if "cisco_nxos" in sshdevicetype.lower():
			fsmshowinvurl = "placeholder"
		fsmtemplatename = sshdevicetype.lower() + '_fsmshowinventory.fsm'
		if not os.path.isfile(fsmtemplatename):
			urllib.urlretrieve(fsmshowinvurl, fsmtemplatename)
		fsmtemplatenamefile = open(fsmtemplatename)
		fsminvtemplate = textfsm.TextFSM(fsmtemplatenamefile)
		tempfilelist.append(fsmtemplatename)
		fsmtemplatenamefile.close()
		# IP Arp Table
		if "cisco_ios" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_iparp.template"
		if "cisco_xe" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_iparp.template"
		if "cisco_nxos" in sshdevicetype.lower():
			fsmshowurl = "placeholder"
		fsmtemplatename = sshdevicetype.lower() + '_fsmiparptable.fsm'
		if not os.path.isfile(fsmtemplatename):
			urllib.urlretrieve(fsmshowurl, fsmtemplatename)
		fsmtemplatenamefile = open(fsmtemplatename)
		fsmarptemplate = textfsm.TextFSM(fsmtemplatenamefile)
		tempfilelist.append(fsmtemplatename)
		fsmtemplatenamefile.close()
		# Mac Table Lists
		if "cisco_ios" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_mac.template"
		if "cisco_xe" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_mac.template"
		if "cisco_nxos" in sshdevicetype.lower():
			fsmshowurl = "placeholder"
		fsmtemplatename = sshdevicetype.lower() + '_fsmmactable.fsm'
		if not os.path.isfile(fsmtemplatename):
			urllib.urlretrieve(fsmshowurl, fsmtemplatename)
		fsmtemplatenamefile = open(fsmtemplatename)
		fsmmactemplate = textfsm.TextFSM(fsmtemplatenamefile)
		tempfilelist.append(fsmtemplatename)
		fsmtemplatenamefile.close()
		# Show Version
		if "cisco_ios" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_version.template"
		if "cisco_xe" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_version.template"
		if "cisco_nxos" in sshdevicetype.lower():
			fsmshowurl = "placeholder"
		fsmtemplatename = sshdevicetype.lower() + '_fsmversion.fsm'
		if not os.path.isfile(fsmtemplatename):
			urllib.urlretrieve(fsmshowurl, fsmtemplatename)
		fsmtemplatenamefile = open(fsmtemplatename)
		fsmvertemplate = textfsm.TextFSM(fsmtemplatenamefile)
		tempfilelist.append(fsmtemplatenamefile)
		fsmtemplatenamefile.close()
		# Show Interface Status
		if "cisco_ios" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_interface_stat.template"
		if "cisco_xe" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_interface_stat.template"
		if "cisco_nxos" in sshdevicetype.lower():
			fsmshowurl = "placeholder"
		fsmtemplatename = sshdevicetype.lower() + '_fsminterfacestat.fsm'
		if not os.path.isfile(fsmtemplatename):
			urllib.urlretrieve(fsmshowurl, fsmtemplatename)
		fsmtemplatenamefile = open(fsmtemplatename)
		fsmintstattemplate = textfsm.TextFSM(fsmtemplatenamefile)
		tempfilelist.append(fsmtemplatenamefile)
		fsmtemplatenamefile.close()
		# Show Power Inline
		if "cisco_ios" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_powerinline.template"
		if "cisco_xe" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_powerinline.template"
		if "cisco_nxos" in sshdevicetype.lower():
			fsmshowurl = "placeholder"
		fsmtemplatename = sshdevicetype.lower() + '_fsmpowerinline.fsm'
		if not os.path.isfile(fsmtemplatename):
			urllib.urlretrieve(fsmshowurl, fsmtemplatename)
		fsmtemplatenamefile = open(fsmtemplatename)
		fsmpoeporttemplate = textfsm.TextFSM(fsmtemplatenamefile)
		tempfilelist.append(fsmtemplatenamefile)
		fsmtemplatenamefile.close()
		# Show IP Interface Brief
		if "cisco_ios" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_powerinline.template"
		if "cisco_xe" in sshdevicetype.lower():
			fsmshowurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_powerinline.template"
		if "cisco_nxos" in sshdevicetype.lower():
			fsmshowurl = "placeholder"
		fsmtemplatename = sshdevicetype.lower() + '_fsmipintbr.fsm'
		if not os.path.isfile(fsmtemplatename):
			urllib.urlretrieve(fsmshowurl, fsmtemplatename)
		fsmtemplatenamefile = open(fsmtemplatename)
		fsmipintbrtemplate = textfsm.TextFSM(fsmtemplatenamefile)
		tempfilelist.append(fsmtemplatenamefile)
		print fsmtemplatenamefile
		fsmtemplatenamefile.close()
		#################################################################
		##################### DOWNLOAD TEMPLATES END ####################
		#################################################################
		#
		#
		#
		#
		#
		#################################################################
		###################### STANDARD ALL START #######################
		#################################################################
		######################## Show Running Config #########################
		sshcommand = showrun
		sshresult = sshnet_connect.send_command(sshcommand)
		showrunresult = sshresult
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		######################## Show Startup Config #########################
		sshcommand = showstart
		sshresult = sshnet_connect.send_command(sshcommand)
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		######################## Show CDP Neighbors #########################
		sshcommand = showcdp
		sshresult = sshnet_connect.send_command(sshcommand)
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		######################## Show LLDP Neighbors #########################
		sshcommand = showlldp
		sshresult = sshnet_connect.send_command(sshcommand)
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)		
		######################## Show Version #########################
		sshcommand = showver
		sshresult = sshnet_connect.send_command(sshcommand)
		#### Find Type of Device ####
		if any(word in sshresult for word in switchlist):
			deviceswitch = 1
		if any(word in sshresult for word in routerlist):
			devicerouter = 1
		if any(word in sshresult for word in fwlist):
			devicefw = 1
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		# Export Version, used later in the full inventory list (NOT A GLOBAL LIST) #
		data = fsmvertemplate.ParseText(sshresult)
		tempversioninfo = []
		for subrow in data:
			# Get Version Number and attach to temporary dictionary
			ver_ver = subrow[0]
			ver_host = subrow[2]
			# Create Temp Dictionary
			tempdict = {}
			# Append Data to Temp Dictionary
			tempdict['Version'] = ver_ver
			tempdict['Hostname'] = ver_host
			# Append Temp Dictionary to Global List
			tempversioninfo.append(tempdict)
		######################## Show Location #########################
		sshcommand = showlocation
		sshresult = sshnet_connect.send_command(sshcommand)
		inv_location = DEF_REMOVEPREFIX(sshresult,'snmp-server location ')
		######################## Show Inventory #########################
		sshcommand = showinv
		sshresult = sshnet_connect.send_command(sshcommand)
		# Find Type of Device #
		if any(word in sshresult for word in switchlist):
			deviceswitch = 1
		if any(word in sshresult for word in routerlist):
			devicerouter = 1
		if any(word in sshresult for word in fwlist):
			devicefw = 1
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		# Export Inventory #
		data = fsminvtemplate.ParseText(sshresult)
		for subrow in data:
			# Get Product Name, Product Serial Number, Description and Stack
			inv_pid = subrow[2]
			inv_sn = subrow[4]
			if re.match('^[1-8]$',subrow[0]) or re.match('^Switch [1-8]$',subrow[0]):
				inv_stack = subrow[0]
				inv_desc = 'Switch chassis'
			else:
				inv_stack = ''
				inv_desc = subrow[0]
			if re.match('^GLC|SFP.*',inv_pid):
				inv_desc = subrow[1]
			# Get Version number from already created list
			for subrow1 in tempversioninfo:
				if sshdevicehostname == subrow1.get('Hostname'):
					inv_ver = subrow1.get('Version')
			# Create Temp Dictionary
			tempdict = {}
			# Append Data to Temp Dictionary
			tempdict['Hostname'] = sshdevicehostname
			tempdict['Product ID'] = inv_pid
			tempdict['Serial Number'] = inv_sn
			tempdict['Description'] = inv_desc
			tempdict['Stack Number'] = inv_stack
			tempdict['Version'] = inv_ver
			tempdict['Location'] = inv_location
			# Append Temp Dictionary to Global List
			fullinventorylist.append(tempdict)
		#################################################################
		######################## STANDARD ALL END #######################
		#################################################################
		#
		#
		#
		#
		#
		#################################################################
		##################### SWITCH SPECIFIC START #####################
		#################################################################
		if deviceswitch == 1:
			######################## Show Mac Address #########################
			sshcommand = showmacaddress
			sshresult = sshnet_connect.send_command(sshcommand)
			if 'invalid' in sshresult:
				sshcommand = showmacaddress_older
				sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
			# Export Mac Address Table
			data = fsmmactemplate.ParseText(sshresult)
			# Get MAC Interface Count
			macintcountb = []
			macintcount = []
			for macintrow0 in data:
				macintname = macintrow0[3]
				tempdict = {}
				# Duplicate Detection
				dupdetect = 0
				for subrow2 in macintcountb:
					if subrow2.get('Interface') == macintname:
						dupdetect = 1
						break
				if dupdetect == 0:
					tempdict['Interface'] = macintname
					macintcountb.append(tempdict)
			for macintrow1 in macintcountb:
				maccount = 0
				macint = macintrow1.get('Interface')
				for subrow2 in data:
					if subrow2[3] == macintrow1.get('Interface'):
						maccount = maccount + 1
				tempdict = {}
				tempdict['Count'] = maccount
				tempdict['Interface'] = macint
				macintcount.append(tempdict)
			# Get MAC addresses and append mac count
			for macintrow2 in data:	
				# Get Hostname, MAC, VLAN, Interface, Count
				mac_host = sshdevicehostname
				mac_mac = macintrow2[0]
				mac_vlan = macintrow2[2]
				mac_int = macintrow2[3]
				for subrow2 in macintcount:
					if mac_int == subrow2.get('Interface'):
						mac_count = subrow2.get('Count')
				# Create Temp Dictionary
				tempdict = {}
				# Append Data to Temp Dictionary
				tempdict['Hostname'] = mac_host
				tempdict['MAC'] = mac_mac
				tempdict['VLAN'] = mac_vlan
				tempdict['Interface'] = mac_int
				tempdict['Count'] = mac_count
				mactablelist.append(tempdict)
			######################## Show Power Budget #########################
			if powerbudget == 1:
				######################## Show Power Inline #########################
				sshcommand = showpowerinline
				sshresult = sshnet_connect.send_command(sshcommand)
				if not 'invalid' in sshresult:
						DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
						# Export Power Inline
						data = fsmpoeporttemplate.ParseText(sshresult)
						for subrow in data:
							# Get Int, Admin, Oper, Power, Device, Class, Max POE
							pow_oper = subrow[2]
							if 'on' in pow_oper.lower():
								pow_oper = 'Up'
							else:
								pow_oper = 'Down'
							# Create Temp Dictionary
							tempdict = {}
							# Append Data to Temp Dictionary
							tempdict['Hostname'] = sshdevicehostname
							tempdict['Interface'] = subrow[0]
							tempdict['Admin Status'] = subrow[1]
							tempdict['Up/Down'] = pow_oper
							tempdict['Power Usage'] = subrow[3]
							tempdict['Device Name'] = subrow[4]
							tempdict['Device Class'] = subrow[5]
							tempdict['Max POE Capability'] = subrow[6]
							# Append Temp Dictionary to Global List
							poeinterfacelist.append(tempdict)
				######################## Show Stack Power #########################
				sshcommand = showpowerinline
				sshresult = sshnet_connect.send_command(showstackpower)
				if not 'invalid' in sshresult:
						DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
			#Show Switch Stack
			sshcommand = showswitch
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)			
			#Show DHCP Snooping
			if dhcpsnooping == 1:
				sshcommand = showdhcpsnooping
				sshresult = sshnet_connect.send_command(sshcommand)
				if not 'invalid' in sshresult:
					DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
			#Show VLAN
			sshcommand = showvlan
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)		
			#Show Trunk
			sshcommand = showtrunk
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
			#Show Spanning-Tree
			sshcommand = showspanning
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
			#Show Spanning-Tree Blocked
			sshcommand = showspanningblock
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		#################################################################
		##################### SWITCH SPECIFIC END #######################
		#################################################################
		#
		#
		#
		#
		#
		#################################################################
		#################### ROUTING SPECIFIC START #####################
		#################################################################
		if 'ip routing' in showrunresult.lower():
			######################## Show IP ARP #########################
			sshcommand = showiparp
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
			# Export ARP Data
			data = fsmarptemplate.ParseText(sshresult)
			for subrow in data:
				# Get IP, MAC, and Interface
				arp_ip = subrow[0]
				# Dup Detect
				dupdetect = 0
				for duprow in ipmactablelist:
					duprowip = duprow.get('IP Address')
					if duprowip == arp_ip:
						dupdetect = 1
						break
				# If no duplicate, append to dictionary/list
				if dupdetect == 1:
					pass
				else:
					arp_age = subrow[1]
					arp_mac = subrow[2]
					arp_host = sshdevicehostname
					arp_int = subrow[4]
					# Create Temp Dictionary
					tempdict = {}
					# Append Data to Temp Dictionary
					tempdict['IP Address'] = arp_ip
					tempdict['MAC'] = arp_mac
					tempdict['Age'] = arp_age
					tempdict['Hostname'] = arp_host
					tempdict['Interface'] = arp_int
					# Append Temp Dictionary to Global List
					ipmactablelist.append(tempdict)
			######################## Show Route Table #########################
			sshcommand = showiproute
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
			if 'router eigrp' in showrunresult.lower():
				#Show EIGRP Neighbors
				sshcommand = showeigrpnei
				sshresult = sshnet_connect.send_command(sshcommand)
				if not 'invalid' in sshresult:
					DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
				#Show EIGRP Topology
				sshcommand = showeigrptop
				sshresult = sshnet_connect.send_command(sshcommand)
				if not 'invalid' in sshresult:
					DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
			if 'router ospf' in showrunresult.lower():
				#Show OSPF Neighbors
				sshcommand = showospfnei
				sshresult = sshnet_connect.send_command(sshcommand)
				if not 'invalid' in sshresult:
					DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
				#Show OSPF Database
				sshcommand = showospfdata
				sshresult = sshnet_connect.send_command(sshcommand)
				if not 'invalid' in sshresult:
					DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
			if 'router bgp' in showrunresult.lower():
				#Show BGP Neighbors
				sshcommand = showbgpnei
				sshresult = sshnet_connect.send_command(sshcommand)
				if not 'invalid' in sshresult:
					DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
				#Show BGP Table
				sshcommand = showbgptable
				sshresult = sshnet_connect.send_command(sshcommand)
				if not 'invalid' in sshresult:
					DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
		if 'multicast-routing' in showrunresult.lower():
			#Show PIM Neighbors
			sshcommand = showippimnei
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
			#Show MRoutes
			sshcommand = showmroute
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
		#################################################################
		#################### ROUTING SPECIFIC END #######################
		#################################################################
		#
		#
		#
		#
		#
		#################################################################
		########################## MISC START ###########################
		#################################################################
		#
		######################## Show Interface Statistics #########################
		sshcommand = showinterfacestat
		sshresult = sshnet_connect.send_command(sshcommand)
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		# Export Interface Statistics #
		data = fsmintstattemplate.ParseText(sshresult)
		for subrow in data:
			# Get Interface,Description,Status,VLAN,Duplex,Speed,Type
			# Create Temp Dictionary
			tempdict = {}
			# Append Data to Temp Dictionary
			tempdict['Hostname'] = sshdevicehostname
			tempdict['Interface'] = subrow[0]
			tempdict['Description'] = subrow[1]
			tempdict['Status'] = subrow[2]
			tempdict['VLAN'] = subrow[3]
			tempdict['Duplex'] = subrow[4]
			tempdict['Speed'] = subrow[5]
			tempdict['Type'] = subrow[6]
			# Append Temp Dictionary to Global List
			l2interfacelist.append(tempdict)
		######################## Show IP Interface Brief #########################
		sshcommand = showipintbr
		sshresult = sshnet_connect.send_command(sshcommand)
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		data = fsmipintbrtemplate.ParseText(sshresult)
		for subrow in data:
			# Get Interface,Description,Status,VLAN,Duplex,Speed,Type
			# Create Temp Dictionary
			tempdict = {}
			# Append Data to Temp Dictionary
			tempdict['Hostname'] = sshdevicehostname
			tempdict['Interface'] = subrow[0]
			tempdict['IP Address'] = subrow[1]
			tempdict['Status'] = subrow[2]
			tempdict['Line Protocol'] = subrow[3]
			# Append Temp Dictionary to Global List
			l3interfacelist.append(tempdict)
		######################## Show IGMP Snooping #########################
		sshcommand = showigmpsnoop
		sshresult = sshnet_connect.send_command(sshcommand)
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
		#Show IGMP Membership
		sshcommand = showipigmpmember
		sshresult = sshnet_connect.send_command(sshcommand)
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
		#Show VRF
		sshcommand = showvrf
		sshresult = sshnet_connect.send_command(sshcommand)
		if not 'invalid' in sshresult:
			DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)	
		#Show Temperature
		if temperature == 1:
			sshcommand = showtemp
			sshresult = sshnet_connect.send_command(sshcommand)
			if not 'invalid' in sshresult:
				DEF_WRITEOUTPUT (sshcommand,sshresult,sshdevicehostname,outputfolder)
		#################################################################
		########################### MISC END ############################
		#################################################################
		sshnet_connect.disconnect()
	except IndexError:
		print 'Could not connect to device ' + sshdeviceip
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except Exception as e:
		print 'Error while gather data with ' + sshdeviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting script'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	print 'Completed device information gathering for ' + sshdeviceip

def DEF_HEALTHCHECK(sshdevice):
	sshdeviceip = sshdevice.get('Device IPs').encode('utf-8')
	sshdevicevendor = sshdevice.get('Vendor').encode('utf-8')
	sshdevicetype = sshdevice.get('Type').encode('utf-8')
	sshdevicetype = sshdevicevendor.lower() + "_" + sshdevicetype.lower()
	### FSM Templates ###
	# FSM Show Interface
	if "cisco_ios" in sshdevicetype:
		fsmshowinturl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_interfaces_health.template"
	if "cisco_xe" in sshdevicetype:
		fsmshowinturl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_interfaces_health.template"
	if "cisco_nxos" in sshdevicetype:
		fsmshowinturl = "placeholder"
	fsmtemplatename = sshdevicetype + '_fsmshowint.fsm'
	if not os.path.isfile(fsmtemplatename):
		urllib.urlretrieve(fsmshowinturl, fsmtemplatename)
	fsmtemplatenamefile = open(fsmtemplatename)
	fsminttemplate = textfsm.TextFSM(fsmtemplatenamefile)
	tempfilelist.append(fsmtemplatename)
	fsmtemplatenamefile.close()
	# FSM Show Temperature
	if "cisco_ios" in sshdevicetype:
		fsmshowtempurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_temp_health.template"
	if "cisco_xe" in sshdevicetype:
		fsmshowtempurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_temp_health.template"
	if "cisco_nxos" in sshdevicetype:
		fsmshowtempurl = "placeholder"	
	fsmtemplatename = sshdevicetype + '_fsmshowtemp.fsm'
	if not os.path.isfile(fsmtemplatename):
		urllib.urlretrieve(fsmshowtempurl, fsmtemplatename)
	fsmtemplatenamefile = open(fsmtemplatename)
	fsmtemptemplate = textfsm.TextFSM(fsmtemplatenamefile)
	tempfilelist.append(fsmtemplatename)
	fsmtemplatenamefile.close()
	#Start Connection
	try:
		for username in usernamelist:
			try:
				try:
					sshusername = username.get('sshusername').encode('utf-8')
					sshpassword = username.get('sshpassword').encode('utf-8')
					enablesecret = username.get('enablesecret').encode('utf-8')
				except:
					sshusername = username.get('sshusername')
					sshpassword = username.get('sshpassword')
					enablesecret = username.get('enablesecret')
				try:
					sshnet_connect = ConnectHandler(device_type=sshdevicetype, ip=sshdeviceip, username=sshusername, password=sshpassword, secret=enablesecret)
					break
				except:
					pass
			except:
				pass
		if not sshnet_connect:
			thread.exit()
		sshdevicehostname = sshnet_connect.find_prompt()
		sshdevicehostname = sshdevicehostname.strip('#')
		if '>' in sshdevicehostname:
			sshnet_connect.enable()
			sshdevicehostname = sshdevicehostname.strip('>')
			sshdevicehostname = sshnet_connect.find_prompt()
			sshdevicehostname = sshdevicehostname.strip('#')
		print 'Health Check starting on ' + sshdevicehostname
		#Show Interfaces
		sshcommand = showinterface
		sshresult = sshnet_connect.send_command(sshcommand)
		hcshowint = fsminttemplate.ParseText(sshresult)
		#Parse through each interface looking for issues
		healthcheckcsv = []
		for hcshowintsingle in hcshowint:
			hcinterfacename = hcshowintsingle[0].encode('utf-8')
			if not 'notconnect' in hcshowintsingle[2]:
				# Look for duplexing issues
				if 'Half-duplex' in hcshowintsingle[6]:
					hcerror = 'Duplex Mismatch'
					hcdescription = hcinterfacename + ' is showing as half-duplex. If this is by design please ignore.'
					healthcheckcsv.append ((sshdevicehostname + ',' + hcerror + ',' + hcdescription))
				if '10Mb/s' in hcshowintsingle[7]:
					hcerror = 'Duplex Mismatch'
					hcdescription = hcinterfacename + ' is showing as 10Mb/s. If this is by design please ignore.'
					healthcheckcsv.append ((sshdevicehostname + ',' + hcerror + ',' + hcdescription))
				# Look for interface counter errors
				# Input Errors
				hcshowintsingleint = hcshowintsingle[8]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)
				if hcshowintsingleint > 0:
					hcerror = 'Input Errors'
					hcinterfacecounter = hcshowintsingle[8]
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' input errors. Usually indicative of a bad link (cabling and/or optic failure).'
					healthcheckcsv.append ((sshdevicehostname + ',' + hcerror + ',' + hcdescription))
				# CRC errors
				hcshowintsingleint = hcshowintsingle[9]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)			
				if hcshowintsingleint > 0:
					hcerror = 'CRC Errors'
					hcinterfacecounter = hcshowintsingle[9]
					hcinterfacecounter = hcinterfacecounter
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' CRC errors. Usually indicative of incorrect duplexing settings or a bad link (cabling and/or optic failure).'
					healthcheckcsv.append ((sshdevicehostname + ',' + hcerror + ',' + hcdescription))
				# Output errors
				hcshowintsingleint = hcshowintsingle[10]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)		
				if hcshowintsingleint > 10000:
					hcerror = 'Saturated Link'
					hcinterfacecounter = hcshowintsingle[10]
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' output errors. This is usually indicative of a saturated interface.  '
					healthcheckcsv.append ((sshdevicehostname + ',' + hcerror + ',' + hcdescription))
				# Collisions
				hcshowintsingleint = hcshowintsingle[11]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)
				if hcshowintsingleint > 0:
					hcerror = 'Shared Medium'
					hcinterfacecounter = hcshowintsingle[11]
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' collisions.  '
					healthcheckcsv.append ((sshdevicehostname + ',' + hcerror + ',' + hcdescription))		
				# Interface resets
				hcshowintsingleint = hcshowintsingle[12]
				if hcshowintsingleint == '':
					hcshowintsingleint = 0
				hcshowintsingleint = int(hcshowintsingleint)			
				if hcshowintsingleint > 20:
					hcerror = 'Interface Reset Count'
					hcinterfacecounter = hcshowintsingle[12]
					hcinterfacecounter = hcinterfacecounter.encode('utf-8')
					hcdescription = hcinterfacename + ' is showing ' + hcinterfacecounter + ' interface resets. '
					healthcheckcsv.append ((sshdevicehostname + ',' + hcerror + ',' + hcdescription))
		#Show Temperature
		try:
			sshcommand = showtemp
			sshresult = sshnet_connect.send_command(sshcommand)
			hcshowtemp = fsmtemptemplate.ParseText(sshresult)
			hctempdegrees = hcshowtemp[0]
			hctempdegrees = hctempdegrees[0]
			hctempdegrees = hctempdegrees.encode('utf-8')
			hctempdegreesint = int(hctempdegrees)
			if hctempdegreesint > 45:
				hcerror = 'Temperature Alert'
				hcdescription = 'Temperature has been recorded at ' + hctempdegrees + ' Celsius. Please lower the temperature for the surrounding environment '
				healthcheckcsv.append ((sshdevicehostname + ',' + hcerror + ',' + hcdescription))
		except:
			pass
		# Exit SSH
		sshnet_connect.disconnect()
		# Parse list into dictionary/list
		saveresultslistsplit = []
		for saveresultsrow in healthcheckcsv:
			saveresultslistsplit.append(saveresultsrow.strip().split(','))
		saveresultslistsplit = [saveresultslistsplit[i:i+3] for i in range(0,len(saveresultslistsplit),3)]
		for saveresultsplitrow in saveresultslistsplit:
			for saveresultssplitrow2 in saveresultsplitrow:
				tempdict = {}
				tempdict['Hostname'] = saveresultssplitrow2[:1][0]
				tempdict['Error'] = saveresultssplitrow2[1:][0]
				tempdict['Description'] = saveresultssplitrow2[2:][0]
				healthchecklist.append(tempdict)
	except IndexError:
		print 'Could not connect to device ' + sshdeviceip
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except Exception as e:
		print 'Error while running health check with ' + sshdeviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting script'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	print 'Completed health check for ' + sshdeviceip
	try:
		sshnet_connect.disconnect()
		thread.exit()
	except:
		pass

def DEF_CDPDISCOVERY(sshusername,sshpassword,enablesecret,cdpseedv,cdpdevicetypev,cdpdiscoverydepthv):
	# Create Commands
	showcdp = "show cdp neighbor detail"
	# FSM Templates
	if "cisco_ios" in cdpdevicetypev.lower():
		fsmshowcdpurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_cdp_nei_detail.template"
	if "cisco_xe" in cdpdevicetypev.lower():
		fsmshowcdpurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_cdp_nei_detail.template"
	if "cisco_nxos" in cdpdevicetypev.lower():
		fsmshowcdpurl = "placeholder"
	fsmtemplatename = cdpdevicetypev.lower() + '_fsmshowcdp.fsm'
	if not os.path.isfile(fsmtemplatename):
		urllib.urlretrieve(fsmshowcdpurl, fsmtemplatename)
	fsmtemplatenamefile = open(fsmtemplatename)
	fsmcdptemplate = textfsm.TextFSM(fsmtemplatenamefile)
	tempfilelist.append(fsmtemplatenamefile)
	fsmtemplatenamefile.close()
	# First Level of Discovery and building the initial seed discovery
	for username in usernamelist:
		try:
			sshusername = username.get('sshusername').encode('utf-8')
			sshpassword = username.get('sshpassword').encode('utf-8')
			enablesecret = username.get('enablesecret').encode('utf-8')
		except:
			sshusername = username.get('sshusername')
			sshpassword = username.get('sshpassword')
			enablesecret = username.get('enablesecret')
		try:
			sshnet_connect = ConnectHandler(device_type=cdpdevicetypev, ip=cdpseedv, username=sshusername, password=sshpassword, secret=enablesecret)
			break
		except:
			pass
	if not sshnet_connect:
		thread.exit()
	sshdevicehostname = sshnet_connect.find_prompt()
	sshdevicehostname = sshdevicehostname.strip('#')
	if '>' in sshdevicehostname:
		sshnet_connect.enable()
		sshdevicehostname = sshdevicehostname.strip('>')
		sshdevicehostname = sshnet_connect.find_prompt()
		sshdevicehostname = sshdevicehostname.strip('#')
	print 'CDP discovery starting on seed device ' + sshdevicehostname
	sshcommand = showcdp
	sshresult = sshnet_connect.send_command(sshcommand)
	hcshowcdp = fsmcdptemplate.ParseText(sshresult)
	print 'Attempting discovery on the seed router'
	cdpdevicediscovery.append(cdpseedv.decode('utf-8'))
	for cdpnei in hcshowcdp:
		try:	
			cdpalreadyexists = 0
			cdpnexthop = 0
			cdpdevicedict = {}
			cdpneiname = cdpnei[0]
			cdpneiip = cdpnei[1]
			cdpneidevice = cdpnei[2]
			cdpneiosfull = cdpnei[5]
			if 'cisco' in cdpneidevice.lower():
				cdpneivend = 'cisco'
				if 'xe' in cdpneiosfull.lower():
					cdpneios = 'xe'
					cdpnexthop = 1
				if 'ios' in cdpneiosfull.lower() and not 'xe' in cdpneiosfull.lower():
					cdpneios = 'ios'
					cdpnexthop = 1
				if 'nxos' in cdpneiosfull.lower() or 'nexus' in cdpneiosfull.lower():
					cdpneios = 'nxos'
					cdpnexthop = 1
				for cdpdevice in cdpdevicecomplete:
					cdpdeviceip = cdpdevice.get('Device IPs').encode('utf-8')
					if cdpdeviceip == cdpneiip:
						cdpalreadyexists = 1
				if cdpalreadyexists == 0 and cdpnexthop == 1:
					cdpdevicedict['Device IPs'] = cdpneiip.decode('utf-8')
					cdpdevicedict['Vendor'] = cdpneivend.decode('utf-8')
					cdpdevicedict['Type'] = cdpneios.decode('utf-8')
					cdpdevicecomplete.append(cdpdevicedict)
		except IndexError:
			print 'Could not connect to device ' + sshdeviceip
			try:
				sshnet_connect.disconnect()
			except:
				'''Nothing'''
		except Exception as e:
			print 'Error while gather data with ' + cdpseedv + '. Error is ' + str(e)
			try:
				sshnet_connect.disconnect()
			except:
				'''Nothing'''
		except KeyboardInterrupt:
			print 'CTRL-C pressed, exiting script'
			try:
				sshnet_connect.disconnect()
			except:
				'''Nothing'''
		
	# Attempt Subsequent Discovery Levels (Non-Threaded)
	def DEF_CDPDISCOVERYSUB(sshusername,sshpassword,enablesecret,cdpip,cdpvendor,cdptype):
		try:
			# FSM Templates
			cdpdevicetype = cdpvendor.lower() + '_' + cdptype.lower()
			if "cisco_ios" in cdpdevicetype.lower():
				fsmshowcdpurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_cdp_nei_detail.template"
			if "cisco_xe" in cdpdevicetype.lower():
				fsmshowcdpurl = "https://raw.githubusercontent.com/routeallthings/Network-Documentation-Automation/master/templates/cisco_ios_show_cdp_nei_detail.template"
			if "cisco_nxos" in cdpdevicetype.lower():
				fsmshowcdpurl = "placeholder"
			fsmtemplatename = cdpdevicetype.lower() + '_fsmshowcdp.fsm'
			if not os.path.isfile(fsmtemplatename):
				urllib.urlretrieve(fsmshowcdpurl, fsmtemplatename)
			fsmtemplatenamefile = open(fsmtemplatename)
			fsmcdptemplate = textfsm.TextFSM(fsmtemplatenamefile)
			tempfilelist.append(fsmtemplatenamefile)
			fsmtemplatenamefile.close()
			# CDP Check
			for username in usernamelist:
				try:
					sshusername = username.get('sshusername').encode('utf-8')
					sshpassword = username.get('sshpassword').encode('utf-8')
					enablesecret = username.get('enablesecret').encode('utf-8')
				except:
					sshusername = username.get('sshusername')
					sshpassword = username.get('sshpassword')
					enablesecret = username.get('enablesecret')
				try:
					sshnet_connect = ConnectHandler(device_type=cdpdevicetype, ip=cdpip, username=sshusername, password=sshpassword, secret=enablesecret)
					break
				except:
					pass
			if not sshnet_connect:
				thread.exit()
			sshdevicehostname = sshnet_connect.find_prompt()
			sshdevicehostname = sshdevicehostname.strip('#')
			if '>' in sshdevicehostname:
				sshnet_connect.enable()
				sshdevicehostname = sshdevicehostname.strip('>')
				sshdevicehostname = sshnet_connect.find_prompt()
				sshdevicehostname = sshdevicehostname.strip('#')
			print 'CDP discovery starting on secondary device ' + sshdevicehostname
			#Show Interfaces
			sshcommand = showcdp
			sshresult = sshnet_connect.send_command(sshcommand)
			hcshowcdp = fsmcdptemplate.ParseText(sshresult)
			for cdpnei in hcshowcdp:
				cdpalreadyexists = 0
				cdpnexthop = 0
				cdpdevicedict = {}
				cdpneiname = cdpnei[0]
				cdpneiip = cdpnei[1]
				cdpneidevice = cdpnei[2]
				cdpneiosfull = cdpnei[5]
				if 'cisco' in cdpneidevice.lower():
					cdpneivend = 'cisco'
					if 'xe' in cdpneiosfull.lower():
						cdpneios = 'xe'
						cdpnexthop = 1
					if 'ios' in cdpneiosfull.lower() and not 'xe' in cdpneiosfull.lower():
						cdpneios = 'ios'
						cdpnexthop = 1
					if 'nxos' in cdpneiosfull.lower() or 'nexus' in cdpneiosfull.lower():
						cdpneios = 'nxos'
						cdpnexthop = 1
					for cdpdevice in cdpdevicecomplete:
						cdpdeviceip = cdpdevice.get('Device IPs').encode('utf-8')
						if cdpdeviceip == cdpneiip:
							cdpalreadyexists = 1
					if cdpalreadyexists == 0 and cdpnexthop == 1:
						cdpdevicedict['Device IPs'] = cdpneiip.decode('utf-8')
						cdpdevicedict['Vendor'] = cdpneivend.decode('utf-8')
						cdpdevicedict['Type'] = cdpneios.decode('utf-8')
						cdpdevicecomplete.append(cdpdevicedict)
						print 'Found new device, adding to list'
			cdpdevicediscovery.append(cdpip.decode('utf-8'))
		except IndexError:
			print 'Could not connect to device ' + sshdeviceip
			try:
				sshnet_connect.disconnect()
			except:
				'''Nothing'''
		except Exception as e:
			print 'Error while CDP data with ' + cdpip + '. Error is ' + str(e)
			cdpdevicediscovery.append(cdpip.decode('utf-8'))
			try:
				sshnet_connect.disconnect()
			except:
				'''Nothing'''
		except KeyboardInterrupt:
			print 'CTRL-C pressed, exiting script'
			try:
				sshnet_connect.disconnect()
			except:
				'''Nothing'''
	# Start CDP Discovery
	cdpdiscoverydepthv = 30
	cdpmaxloop = cdpdiscoverydepthv * 3
	cdpmaxloopiteration = 0
	if cdpdevicecomplete:
		while cdpmaxloopiteration < cdpmaxloop:
			for cdpdevice in cdpdevicecomplete:
				cdpalreadyexists=0
				cdpip = cdpdevice.get('Device IPs').encode('utf-8')
				for cdpalreadyattempted in cdpdevicediscovery:
					if cdpip == cdpalreadyattempted:
						cdpalreadyexists = 1
				cdpvendor = cdpdevice.get('Vendor').encode('utf-8')
				cdptype = cdpdevice.get('Type').encode('utf-8')
				if not cdpalreadyexists == 1:
					DEF_CDPDISCOVERYSUB(sshusername,sshpassword,enablesecret,cdpip,cdpvendor,cdptype)
					cdpmaxloopiteration = cdpmaxloopiteration + 1
##### END OF FUNCTIONS #####				


# Start of threading
print '----Starting to gather data from equipment----'
if __name__ == "__main__":
	# Get Devices from XLSX and lowercase it all
	for xlsxdevices in xlhelper.sheet_to_dict(excelfilelocation,'Device IPs'):
		try:
			xlsxdeviceslist = {}
			xlsxdeviceslist['Device IPs'] = xlsxdevices.get('Device IPs').encode('utf-8').lower()
			xlsxdeviceslist['Vendor'] = xlsxdevices.get('Vendor').encode('utf-8').lower()
			xlsxdeviceslist['Type'] = xlsxdevices.get('Type').encode('utf-8').lower()
			sshdevices.append(xlsxdeviceslist)
		except:
			pass
	# Get Devices from CDP and check to ignore duplicates imported from XLSX
	if devicediscoveryv == 1 or not 'na' in deviceseedv.lower():
		# SNMP Section
		graph = mnetsuite.mnet_graph()
		opt_dot = None
		opt_depth = devicediscoverydepthv
		opt_title = devicediscoverymaptitlev
		opt_conf = mnetfile
		opt_catalog = mnetcat
		graph.set_max_depth(opt_depth)
		graph.load_config(mnetfile)
		graph.crawl(devicediscoveryseedv)
		graph.output_catalog(mnetcat)
		# null byte check
		mnetcatfile = open(mnetcat, 'rb')
		mnetcatdata = mnetcatfile.read()
		if not mnetcatdata.find('\x00') == 1:
			mnetcatfilew = open(mnetcat, 'wb')
			mnetcatfilew.write(mnetcatdata.replace('\x00', ''))
			mnetcatfilew.close()
			mnetcatfile.close()
			mnetcatfile = open (mnetcat, 'rb')
		mnetcatalog = csv.reader(mnetcatfile, delimiter=',')
		for row in mnetcatalog:
			try:
				# Point based system to match devices
				ciscopoints = 0
				if row[1] == '':
					ciscopoints = ciscopoints - 50
				if row[6] != 'None':
					ciscopoints = ciscopoints + 5
				if re.match(ciscoverreg, row[3]):
					ciscopoints = ciscopoints + 5
				# Check point total for Cisco match
				if ciscopoints > 4:
					snmpdiscoverylist = {}
					snmpdeviceip = snmpdiscoverylist['Device IPs'] = row[1]
					snmpdiscoverylist['Device IPs'] = snmpdeviceip
					snmpdiscoverylist['Vendor'] = 'Cisco'
					# Check for device type logic
					if any(word in row[2] for word in ciscoxelist):
						snmpdiscoverydevicetype = 'xe'
					if any(word in row[2] for word in ciscoioslist):
						snmpdiscoverydevicetype = 'ios'
					if any(word in row[2] for word in cisconxoslist):
						snmpdiscoverydevicetype = 'nxos'
					snmpdiscoverylist['Type'] = snmpdiscoverydevicetype
					snmpduplicate = 0
					for sshdevice in sshdevices:
						try:
							if snmpdeviceip == sshdevice.get('Device IPs').encode('utf-8'):
								snmpduplicate = 1
						except:
							pass
					if snmpduplicate == 0:
						sshdevices.append(snmpdiscoverylist)
			except:
				pass
		mnetcatfile.close()
		# SSH Section
		if devicediscoverysshv == 1:
			print 'Starting SSH CDP Discovery'
			DEF_CDPDISCOVERY(usernamelist,devicediscoveryseedv,devicediscoverytypev,devicediscoverydepthv)
			if cdpdevicecomplete:
				for cdpdevice in cdpdevicecomplete:
					cdpduplicate = 0
					cdpdeviceip = cdpdevice.get('Device IPs').encode('utf-8')
					for sshdevice in sshdevices:
						try:
							if cdpdeviceip == sshdevice.get('Device IPs').encode('utf-8'):
								cdpduplicate = 1
						except:
							pass
					if cdpduplicate == 0:
						sshdevices.append(cdpdevice)
	# Start Threads
	for sshdevice in sshdevices:	
		sshdeviceip = sshdevice.get('Device IPs').encode('utf-8')
		print "Spawning Thread for " + sshdeviceip
		t = threading.Thread(target=DEF_STARTALLTESTS, args=(sshdevice,))
		t.start()
	main_thread = threading.currentThread()
	for it_thread in threading.enumerate():
		if it_thread != main_thread:
			it_thread.join()

################################ EXPORTING REPORTS ################################
print 'Exporting Informational Reports'
### Full Inventory ###
try:
	wb = Workbook()
	dest_filename = 'Full-Inventory.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	ws1.title = "Device Inventory"
	ws1.append(['Hostname','Product ID','Serial Number','Stack Number','Manufacture Date','Version','Description'])
	startrow = 2
	for row in fullinventorylist:
		if 'chassis' in row.get('Description').lower():
			# Attempt to find the age of the device
			try:
				age_base = 1996
				age_year = int(row.get('Serial Number')[3:5])
				age_week = (row.get('Serial Number')[5:7])
				age_year_manufactured = age_base + age_year
				age_manufactured = datetime.datetime.strptime(str(age_year_manufactured) + '-W' + age_week.encode('utf-8') + '-0', '%Y-W%W-%w')
				age_manufactured = '{:%B %d, %Y}'.format(age_manufactured)
			except:
				age_manufactured = ''
			# Add to workbook
			ws1['A' + str(startrow)] = row.get('Hostname')
			ws1['B' + str(startrow)] = row.get('Product ID')
			ws1['C' + str(startrow)] = row.get('Serial Number')
			ws1['D' + str(startrow)] = row.get('Stack Number')
			ws1['E' + str(startrow)] = age_manufactured
			ws1['F' + str(startrow)] = row.get('Version')
			ws1['G' + str(startrow)] = row.get('Description')
			startrow = startrow + 1
	ws2 = wb.create_sheet(title="Module Inventory")
	ws2.append(['Hostname','Product ID','Serial Number','Description'])
	startrow = 2
	for row in fullinventorylist:
		if not 'chassis' in row.get('Description').lower():
			ws2['A' + str(startrow)] = row.get('Hostname')
			ws2['B' + str(startrow)] = row.get('Product ID')
			ws2['C' + str(startrow)] = row.get('Serial Number')
			ws2['D' + str(startrow)] = row.get('Description')
			startrow = startrow + 1
	wb.save(filename = dest_path)
	print 'Successfully created Full Inventory Report'
except Exception as e:
	print 'Error creating Full Inventory Report. Error is ' + str(e)
#### MAC and ARP Report ###
try:
	wb = Workbook()
	dest_filename = 'ARP-MAC-Report.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	ws1.title = "ARP Report"
	ws1.append(['IP Address','MAC','Manufacturer','Source Device','Inteface','MAC Count on Interface'])
	# Create ARP report by looking for closest hop interface
	skiparpreport = 0
	# Preload MAC DB
	try:
		maclookupfilename = 'oui.txt'
		tempfilelist.append(maclookupfilename)
		if not os.path.isfile(maclookupfilename):
			urllib.urlretrieve(maclookupdburl, maclookupfilename)
		maclookupdbo = open(maclookupfilename)
		maclookupdb = maclookupdbo.readlines()
		maclookupdbo.close()
		skipmac = 0
	except Exception as e:
		skipmac = 1
		print 'Could not load MAC database. Error is ' + str(e)
	# Start processing data
	for row in ipmactablelist:
		tempdict = {}
		tempdict['IP Address'] = row.get('IP Address')
		tempdict['MAC'] = row.get('MAC')
		mac_company_mac = (row.get('MAC')[0:8])
		# Get a vendor mac address and add to the table
		if skipmac == 0:
			try:
				for line in maclookupdb:
					if line.startswith(mac_company_mac):
						maccompany = (re.search(r'^(([A-Z0-9]{2}[-]){2}[A-Z0-9]{2}.*\(hex\)\s+)(.*)',line)).group(3)
					if maccompany == '':
						maccompany = 'Unknown'
			except:
				maccompany = 'Unknown'
		tempdict['MAC Manufacturer'] = maccompany
		if '-' in row.get('Age'):
			tempdict['Source Device'] = row.get('Hostname')
			tempdict['Interface'] = row.get('Interface')
			tempdict['MAC Count on Interface'] = 1
		else:
			# Find the lowest count interface in the list that matches the mac address
			maccount = 100000
			for temprow in mactablelist:
				if temprow.get('Count') <= maccount and temprow.get('MAC') == row.get('MAC'):
					maccount = temprow.get('Count')
					macint = temprow.get('Interface')
					machost = temprow.get('Hostname')
			# Bug Fix - Somehow some interfaces were being missed (showing 100000 mac), finding interface match from earlier and matching on that
			'''
			if maccount == 100000:
				for temprow in mactablelist:
					if macint == temprow.get('Interface'):
						maccount = temprow.get('Count')
			'''
			tempdict['Source Device'] = machost
			tempdict['Interface'] = macint
			tempdict['MAC Count on Interface'] = maccount
		iparptablelist.append(tempdict)
	# Create the actual ARP report
	startrow = 2
	for row in iparptablelist:
		ws1['A' + str(startrow)] = row.get('IP Address')
		ws1['B' + str(startrow)] = row.get('MAC')
		ws1['C' + str(startrow)] = row.get('MAC Manufacturer')
		ws1['D' + str(startrow)] = row.get('Source Device')
		ws1['E' + str(startrow)] = row.get('Interface')
		ws1['F' + str(startrow)] = row.get('MAC Count on Interface')
		startrow = startrow + 1
	# Change worksheet
	ws2 = wb.create_sheet(title="MAC")
	ws2.append(['Hostname','MAC','Manufacturer','VLAN','Interface'])
	startrow = 2
	for row in mactablelist:
		# Get Manufacturer of MAC
		if skipmac == 0:
			try:
				for line in maclookupdb:
					if line.startswith(mac_company_mac):
						maccompany = (re.search(r'^(([A-Z0-9]{2}[-]){2}[A-Z0-9]{2}.*\(hex\)\s+)(.*)',line)).group(3)
					if maccompany == '':
						maccompany = 'Unknown'
			except:
				maccompany = 'Unknown'
		# Append to Workbook
		ws2['A' + str(startrow)] = row.get('Hostname')
		ws2['B' + str(startrow)] = row.get('MAC')
		ws2['C' + str(startrow)] = maccompany
		ws2['D' + str(startrow)] = row.get('VLAN')
		ws2['E' + str(startrow)] = row.get('Interface')
		startrow = startrow + 1
	# Save workbook
	wb.save(filename = dest_path)
	print 'Successfully created ARP/MAC Report'
except Exception as e:
	print 'Could not save the ARP/MAC data to XLSX. Error is ' + str(e)
### Interface Report ###
try:
	wb = Workbook()
	dest_filename = 'Interface-Report.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	ws1.title = "Interface Overview"
	ws1.append(['Hostname','100Mb','1Gb','10gb','40gb','100gb','POE'])
	startrow = 2
	# Populate Device Names
	l2devicenames = []
	for row in l2interfacelist:
		dupdetect = 0	
		for device in l2devicenames:
			if row.get('Hostname') == device:
				dupdetect = 1
		if dupdetect == 0:
			l2devicenames.append(row.get('Hostname'))
	# Count Interfaces
	for row in l2devicenames:
		int_hostname = row
		faint = 0
		geint = 0
		tengeint = 0
		fortygeint = 0
		hundredgeint = 0
		poeint = 0
		for subrow in l2interfacelist:
			if '10/100BaseTX' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				faint = faint + 1
			if '10/100/1000BaseTX' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				geint = geint + 1
			if '10/100/10000BaseTX' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				tengeint = tengeint + 1
			if '10/40GB' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				fortygeint = fortygeint + 1
			if '10/25/50/100' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				hundredgeint = hundredgeint + 1
			if '10/40/100' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				hundredgeint = hundredgeint + 1
			for subrow1 in poeinterfacelist:
				if subrow.get('Interface') == subrow1.get('Interface') and Decimal(subrow1.get('Power Usage')) > 0 and subrow.get('Hostname') == subrow1.get('Hostname'):
					poeint = poeint + 1
		# Get Hostname and populate
		ws1['A' + str(startrow)] = int_hostname
		ws1['B' + str(startrow)] = faint
		ws1['C' + str(startrow)] = geint
		ws1['D' + str(startrow)] = tengeint
		ws1['E' + str(startrow)] = fortygeint
		ws1['F' + str(startrow)] = hundredgeint
		ws1['G' + str(startrow)] = poeint
		startrow = startrow + 1
	try:
		ws2 = wb.create_sheet(title="L2 Interfaces")
		ws2.append(['Hostname','Interface','Type','Status','Speed','Duplex','VLAN','POE'])
		startrow = 2
		for row in l2interfacelist:
			# Get POE interfaces and combine
			for subrow in poeinterfacelist:
				if row.get('Interface') == subrow.get('Interface') and Decimal(subrow.get('Power Usage')) > 0 and row.get('Hostname') == subrow.get('Hostname'):
					poeint = 'Yes'
				else:
					poeint = 'No'
			ws2['A' + str(startrow)] = row.get('Hostname')
			ws2['B' + str(startrow)] = row.get('Interface')
			ws2['C' + str(startrow)] = row.get('Type')
			ws2['D' + str(startrow)] = row.get('Status')
			ws2['E' + str(startrow)] = row.get('Speed')
			ws2['F' + str(startrow)] = row.get('Duplex')
			ws2['G' + str(startrow)] = row.get('VLAN').encode('utf-8')
			ws2['H' + str(startrow)] = poeint
			startrow = startrow + 1
	except Exception as e:
		print 'Error creating L2 Interface Report. Error is ' + str(e)	
	try:
		ws3 = wb.create_sheet(title="L3 Interfaces")
		ws3.append(['Hostname','Interface','IP Address'])
		startrow = 2
		for row in l3interfacelist:
			ws3['A' + str(startrow)] = row.get('Hostname')
			ws3['B' + str(startrow)] = row.get('Interface')
			ws3['C' + str(startrow)] = row.get('IP Address')
			startrow = startrow + 1
			ws3 = wb.create_sheet(title="L3 Interfaces")
	except Exception as e:
		print 'Error creating L3 Interface Report. Error is ' + str(e)	
	wb.save(filename = dest_path)
	print 'Successfully created Interface Report'
except Exception as e:
	print 'Error creating Interface Report. Error is ' + str(e)	
### POE Report ###
try:
	wb = Workbook()
	dest_filename = 'POE-Report.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	ws1.title = "POE Interfaces"
	ws1.append(['Hostname','Interface','Admin Status','Operational Status','Power Usage','Device Name','Device Class','Max POE Capability'])
	startrow = 2
	for row in poeinterfacelist:
		if not row.get('Power Usage') == '0.0':
			# Add to workbook
			ws1['A' + str(startrow)] = row.get('Hostname')
			ws1['B' + str(startrow)] = row.get('Interface')
			ws1['C' + str(startrow)] = row.get('Admin Status')
			ws1['D' + str(startrow)] = row.get('Up/Down')
			ws1['E' + str(startrow)] = row.get('Power Usage')
			ws1['F' + str(startrow)] = row.get('Device Name')
			ws1['G' + str(startrow)] = row.get('Device Class')
			ws1['H' + str(startrow)] = row.get('Max POE Capability')
			startrow = startrow + 1
	wb.save(filename = dest_path)
	print 'Successfully created POE Report'
except Exception as e:
	print 'Error creating POE Report. Error is ' + str(e)	
	
	
### Health Check ###
try:
	if healthcheckv == 1:
		print 'Exporting Health Reports'
		wb = Workbook()
		dest_filename = 'Health-Check-Report.xlsx'
		dest_path = exportlocation + '\\' + dest_filename
		ws1 = wb.active
		ws1.title = "Health Check"
		ws1.append(['Hostname','Error','Description'])
		startrow = 2
		for row in healthchecklist:
			ws1['A' + str(startrow)] = row.get('Hostname')
			ws1['B' + str(startrow)] = row.get('Error')
			ws1['C' + str(startrow)] = row.get('Description')
			startrow = startrow + 1		
		wb.save(filename = dest_path)
		print 'Successfully created Health Check Report'
except Exception as e:
	print 'Could not save health check data to CSV. Error is ' + str(e)
	
# Map Output
print 'Starting to export Map Topology'
try:
	if devicediscoverymapv == 1:
		topologyfile = exportlocation + '\\Network_Topology.svg'
		topologyname = devicediscoverymaptitlev
		graph.output_dot(topologyfile, topologyname)
except Exception as e:
	print 'Error with exporting network topology. Error is ' + str(e)
# Cleanup
print 'Starting Temp File Cleanup'
for file in tempfilelist:
	try:
		os.remove(file.name)
	except:
		pass
		